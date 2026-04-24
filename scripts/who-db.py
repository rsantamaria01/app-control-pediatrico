#!/usr/bin/env python3
"""Build data/who.db from the WHO Child Growth Standards expanded LMS tables.

Downloads 8 .xlsx files from cdn.who.int (boys + girls for 4 indicators),
parses the L, M, S columns from each sheet, and writes a single SQLite
database with one table: who_lms_standards.

Run via ``bash scripts/who-db.sh`` (which manages a venv and dependencies).
Direct invocation: ``python scripts/who-db.py --output data/who.db``.

The output file is committed to the repository and is read-only at runtime.
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
import tempfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook


@dataclass(frozen=True)
class Source:
    indicator: str  # HEIGHT_FOR_AGE | WEIGHT_FOR_AGE | WEIGHT_FOR_HEIGHT | BMI_FOR_AGE
    gender: str  # MALE | FEMALE
    x_column: str  # 'age' for age-based; 'height' for weight-for-height
    label: str
    url: str


SOURCES: tuple[Source, ...] = (
    Source(
        "HEIGHT_FOR_AGE", "MALE", "age", "lhfa-boys",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/length-height-for-age/"
        "expandable-tables/lhfa-boys-zscore-expanded-tables.xlsx",
    ),
    Source(
        "HEIGHT_FOR_AGE", "FEMALE", "age", "lhfa-girls",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/length-height-for-age/"
        "expandable-tables/lhfa-girls-zscore-expanded-tables.xlsx",
    ),
    Source(
        "WEIGHT_FOR_AGE", "MALE", "age", "wfa-boys",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/weight-for-age/"
        "expandable-tables/wfa-boys-zscore-expanded-tables.xlsx",
    ),
    Source(
        "WEIGHT_FOR_AGE", "FEMALE", "age", "wfa-girls",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/weight-for-age/"
        "expandable-tables/wfa-girls-zscore-expanded-tables.xlsx",
    ),
    Source(
        "WEIGHT_FOR_HEIGHT", "MALE", "height", "wfh-boys",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/weight-for-height/"
        "expandable-tables/wfh-boys-zscore-expanded-tables.xlsx",
    ),
    Source(
        "WEIGHT_FOR_HEIGHT", "FEMALE", "height", "wfh-girls",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/weight-for-height/"
        "expandable-tables/wfh-girls-zscore-expanded-tables.xlsx",
    ),
    Source(
        "BMI_FOR_AGE", "MALE", "age", "bfa-boys",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/body-mass-index-for-age/"
        "expandable-tables/bfa-boys-zscore-expanded-tables.xlsx",
    ),
    Source(
        "BMI_FOR_AGE", "FEMALE", "age", "bfa-girls",
        "https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators/body-mass-index-for-age/"
        "expandable-tables/bfa-girls-zscore-expanded-tables.xlsx",
    ),
)


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS who_lms_standards (
  id        INTEGER PRIMARY KEY AUTOINCREMENT,
  indicator TEXT NOT NULL,
  gender    TEXT NOT NULL,
  x_value   REAL NOT NULL,
  l         REAL NOT NULL,
  m         REAL NOT NULL,
  s         REAL NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_who_lms_lookup
  ON who_lms_standards (indicator, gender, x_value);
"""


def fail(msg: str, *, code: int = 1) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(code)


def download(url: str, label: str, idx: int, total: int) -> bytes:
    print(f"[{idx}/{total}] Downloading {label}...", end=" ", flush=True)
    try:
        resp = requests.get(url, timeout=60)
    except requests.RequestException as exc:
        print("FAIL")
        fail(f"network error fetching {url}: {exc}")
    if resp.status_code != 200:
        print("FAIL")
        fail(
            f"unexpected HTTP status {resp.status_code} for {url}\n"
            f"body (first 200 chars): {resp.text[:200]!r}"
        )
    size_kb = len(resp.content) // 1024
    print(f"OK ({size_kb} KB)")
    return resp.content


def parse_xlsx(blob: bytes, source: Source) -> list[tuple[float, float, float, float]]:
    """Return a list of (x_value, L, M, S) tuples parsed from the WHO sheet.

    Locates columns by header name (case-insensitive). The X column is
    'age' (months) for age-based indicators and 'height' (or 'length') for
    weight-for-height. Skips any rows whose X is non-numeric (header/footer).
    """
    workbook = load_workbook(filename=BytesIO(blob), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    header_row: tuple | None = None
    for row in rows:
        if row is None:
            continue
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if not non_empty:
            continue
        # header row contains 'L', 'M', 'S' (case-insensitive) somewhere
        upper = [str(c).strip().lower() for c in row if c is not None]
        if "l" in upper and "m" in upper and "s" in upper:
            header_row = row
            break
    if header_row is None:
        fail(f"could not locate L/M/S header row in {source.label}")

    header_lookup: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        key = str(value).strip().lower()
        if key not in header_lookup:
            header_lookup[key] = idx

    if source.x_column == "age":
        x_keys = ("age", "month", "agemons")
    else:
        x_keys = ("height", "length", "lengthheight")
    x_idx: int | None = None
    for key in x_keys:
        if key in header_lookup:
            x_idx = header_lookup[key]
            break
    if x_idx is None:
        fail(
            f"missing X column ({'/'.join(x_keys)}) in {source.label}; "
            f"saw headers: {sorted(header_lookup)}"
        )

    try:
        l_idx = header_lookup["l"]
        m_idx = header_lookup["m"]
        s_idx = header_lookup["s"]
    except KeyError as exc:
        fail(f"missing column {exc} in {source.label}")

    parsed: list[tuple[float, float, float, float]] = []
    for row in rows:
        if row is None:
            continue
        try:
            x_raw = row[x_idx]
            l_raw = row[l_idx]
            m_raw = row[m_idx]
            s_raw = row[s_idx]
        except IndexError:
            continue
        if x_raw is None or l_raw is None or m_raw is None or s_raw is None:
            continue
        try:
            x = float(x_raw)
            l = float(l_raw)
            m = float(m_raw)
            s = float(s_raw)
        except (TypeError, ValueError):
            continue
        parsed.append((x, l, m, s))

    if not parsed:
        fail(f"no numeric LMS rows parsed from {source.label}")
    return parsed


def write_db(target: Path, all_rows: Iterable[tuple[Source, list[tuple[float, float, float, float]]]]) -> int:
    target.parent.mkdir(parents=True, exist_ok=True)
    # Write to a temp file in the same directory, then atomic rename on success.
    fd, tmp_path = tempfile.mkstemp(prefix="who-", suffix=".db", dir=str(target.parent))
    os.close(fd)
    tmp_db = Path(tmp_path)
    try:
        conn = sqlite3.connect(tmp_db)
        try:
            conn.executescript(SCHEMA_SQL)
            total = 0
            for source, rows in all_rows:
                conn.execute("BEGIN")
                conn.executemany(
                    "INSERT INTO who_lms_standards "
                    "(indicator, gender, x_value, l, m, s) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    [
                        (source.indicator, source.gender, x, l, m, s)
                        for (x, l, m, s) in rows
                    ],
                )
                conn.commit()
                total += len(rows)
            conn.execute("VACUUM")
        finally:
            conn.close()
        os.replace(tmp_db, target)
        return total
    except Exception:
        try:
            tmp_db.unlink()
        except FileNotFoundError:
            pass
        raise


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parent.parent / "data" / "who.db"),
        help="Path for the output SQLite database (default: data/who.db).",
    )
    args = parser.parse_args()
    output_path = Path(args.output).resolve()

    total = len(SOURCES)
    parsed_all: list[tuple[Source, list[tuple[float, float, float, float]]]] = []
    for idx, source in enumerate(SOURCES, start=1):
        blob = download(source.url, source.label, idx, total)
        print(f"[{idx}/{total}] Parsing {source.label}...", end=" ", flush=True)
        rows = parse_xlsx(blob, source)
        print(f"OK ({len(rows)} rows)")
        parsed_all.append((source, rows))

    print(f"Writing {output_path}...")
    total_rows = write_db(output_path, parsed_all)
    print(
        f"✓ {output_path} created — {len(SOURCES)} indicators loaded, "
        f"{total_rows} total rows"
    )


if __name__ == "__main__":
    main()
